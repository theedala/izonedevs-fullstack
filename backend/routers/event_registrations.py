from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from io import BytesIO
from fastapi.responses import StreamingResponse

from database import get_db, EventRegistration, Event, User
from schemas import (
    EventRegistrationCreate,
    EventRegistration as EventRegistrationSchema,
    EventRegistrationWithEvent,
    APIResponse,
    PaginatedResponse
)
from auth import get_current_active_user, require_admin
from services.qr_service import QRCodeService
from services.email_service import EmailService

router = APIRouter()


@router.post("/{event_id}/register", response_model=EventRegistrationSchema)
async def register_for_event(
    event_id: int,
    registration_data: EventRegistrationCreate,
    db: Session = Depends(get_db)
):
    """Register for an event"""
    # Check if event exists and is available for registration
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    if event.status in ["cancelled", "completed"]:
        raise HTTPException(status_code=400, detail="Event is not available for registration")
    
    # Check if user already registered
    existing_registration = db.query(EventRegistration).filter(
        EventRegistration.event_id == event_id,
        EventRegistration.email == registration_data.email
    ).first()
    
    if existing_registration:
        raise HTTPException(status_code=400, detail="You are already registered for this event")
    
    # Check if event is full
    if event.max_attendees:
        current_registrations = db.query(EventRegistration).filter(
            EventRegistration.event_id == event_id,
            EventRegistration.registration_status == "confirmed"
        ).count()
        
        if current_registrations >= event.max_attendees:
            raise HTTPException(status_code=400, detail="Event is full")
    
    # Create registration
    db_registration = EventRegistration(
        event_id=event_id,
        **registration_data.dict()
    )
    db.add(db_registration)
    db.commit()
    db.refresh(db_registration)
    
    # Generate QR code
    qr_file_path = QRCodeService.get_qr_file_path(db_registration.id, event_id)
    qr_data = {
        "id": db_registration.id,
        "event_id": event_id,
        "event_title": event.title,
        "name": db_registration.name,
        "email": db_registration.email,
        "created_at": db_registration.created_at.isoformat(),
        "registration_status": db_registration.registration_status
    }
    
    QRCodeService.generate_qr_code(qr_data, qr_file_path)
    
    # Update registration with QR code path
    db_registration.qr_code_path = qr_file_path
    db.commit()
    
    # Send confirmation email with QR code
    try:
        await EmailService.send_event_registration_confirmation(
            email=db_registration.email,
            name=db_registration.name,
            event=event,
            registration=db_registration,
            qr_code_path=qr_file_path
        )
    except Exception as e:
        print(f"Failed to send confirmation email: {e}")
    
    return db_registration


@router.get("/registrations", response_model=PaginatedResponse)
async def get_event_registrations(
    page: int = Query(1, ge=1),
    size: int = Query(10, ge=1, le=100),
    event_id: Optional[int] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Get event registrations (admin only)"""
    query = db.query(EventRegistration)
    
    if event_id:
        query = query.filter(EventRegistration.event_id == event_id)
    
    if status:
        query = query.filter(EventRegistration.registration_status == status)
    
    if search:
        query = query.filter(
            (EventRegistration.name.contains(search)) |
            (EventRegistration.email.contains(search)) |
            (EventRegistration.organization.contains(search))
        )
    
    total = query.count()
    registrations = query.order_by(EventRegistration.created_at.desc()).offset((page - 1) * size).limit(size).all()
    
    # Include event details
    results = []
    for registration in registrations:
        reg_dict = EventRegistrationSchema.from_orm(registration).dict()
        reg_dict['event'] = {
            'id': registration.event.id,
            'title': registration.event.title,
            'start_date': registration.event.start_date,
            'location': registration.event.location
        }
        results.append(reg_dict)
    
    return PaginatedResponse(
        items=results,
        total=total,
        page=page,
        size=size,
        pages=(total + size - 1) // size
    )


@router.get("/registrations/export")
async def export_registrations_excel(
    event_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Export event registrations to Excel (admin only)"""
    query = db.query(EventRegistration).join(Event)
    
    if event_id:
        query = query.filter(EventRegistration.event_id == event_id)
    
    if status:
        query = query.filter(EventRegistration.registration_status == status)
    
    registrations = query.order_by(EventRegistration.created_at.desc()).all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Event Registrations"
    
    # Define headers
    headers = [
        "Registration ID", "Event Title", "Name", "Email", "Phone", 
        "Organization", "Experience Level", "Interests", 
        "Dietary Restrictions", "Special Requirements", "Status", 
        "Registration Date"
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E348A", end_color="2E348A", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Add data
    for row, registration in enumerate(registrations, 2):
        data = [
            registration.id,
            registration.event.title,
            registration.name,
            registration.email,
            registration.phone or "",
            registration.organization or "",
            registration.experience_level or "",
            registration.interests or "",
            registration.dietary_restrictions or "",
            registration.special_requirements or "",
            registration.registration_status,
            registration.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"event_registrations_{timestamp}.xlsx"
    
    return StreamingResponse(
        BytesIO(excel_buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.put("/registrations/{registration_id}/status", response_model=APIResponse)
async def update_registration_status(
    registration_id: int,
    status: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Update registration status (admin only)"""
    registration = db.query(EventRegistration).filter(EventRegistration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")
    
    valid_statuses = ["confirmed", "cancelled", "attended"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")
    
    registration.registration_status = status
    db.commit()
    
    return APIResponse(
        success=True,
        message=f"Registration status updated to {status}"
    )


@router.delete("/registrations/{registration_id}", response_model=APIResponse)
async def delete_registration(
    registration_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Delete registration (admin only)"""
    registration = db.query(EventRegistration).filter(EventRegistration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")
    
    # Remove QR code file if it exists
    if registration.qr_code_path and os.path.exists(registration.qr_code_path):
        os.remove(registration.qr_code_path)
    
    db.delete(registration)
    db.commit()
    
    return APIResponse(
        success=True,
        message="Registration deleted successfully"
    )